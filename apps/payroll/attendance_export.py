"""Attendance export helpers — Excel (openpyxl) and HTML-for-print."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


STATUS_COLORS = {
    'present':  'D1FAE5',
    'absent':   'FEE2E2',
    'half_day': 'FEF9C3',
    'leave':    'DBEAFE',
    'holiday':  'EDE9FE',
    'week_off': 'F3F4F6',
    'future':   'FFFFFF',
}

STATUS_LABELS = {
    'present':  'P',
    'absent':   'A',
    'half_day': 'H',
    'leave':    'L',
    'holiday':  'HO',
    'week_off': 'WO',
    'future':   '—',
}


def _thin_border():
    side = Side(style='thin', color='D1D5DB')
    return Border(left=side, right=side, top=side, bottom=side)


def export_excel(calendar_data):
    """Returns an openpyxl Workbook for the given calendar_data dict."""
    wb = openpyxl.Workbook()
    ws = wb.active

    emp_name   = calendar_data['employee_name']
    emp_code   = calendar_data['employee_code']
    month_name = calendar_data['month_name']
    year       = calendar_data['year']
    days       = calendar_data['days']
    summary    = calendar_data['summary']

    ws.title = f'{emp_code} {month_name[:3]}'

    # ── Title ──────────────────────────────────────────────
    ws.merge_cells('A1:H1')
    title_cell = ws['A1']
    title_cell.value = f'Attendance Report — {emp_name} ({emp_code}) — {month_name} {year}'
    title_cell.font      = Font(bold=True, size=13)
    title_cell.alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 22

    # ── Summary row ────────────────────────────────────────
    ws.merge_cells('A2:H2')
    summary_text = (
        f"Present: {summary['present']}  |  Absent: {summary['absent']}  |  "
        f"Half Day: {summary['half_day']}  |  Leave: {summary['leave']}  |  "
        f"Holiday: {summary['holiday']}  |  Week Off: {summary['week_off']}  |  "
        f"Paid Days: {summary['paid_days']}/{summary['working_days']}"
    )
    ws['A2'].value     = summary_text
    ws['A2'].font      = Font(size=9, color='6B7280')
    ws['A2'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[2].height = 16

    # ── Header ─────────────────────────────────────────────
    headers = ['Day', 'Date', 'Day', 'Status', 'Status Label', 'Leave Type', 'Notes', 'Marked By']
    header_fill = PatternFill(fill_type='solid', fgColor='1F2937')
    header_font = Font(bold=True, color='FFFFFF', size=10)

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border    = _thin_border()

    ws.row_dimensions[3].height = 18

    # ── Data rows ──────────────────────────────────────────
    for row_idx, day in enumerate(days, start=4):
        status = day['status']
        fill   = PatternFill(fill_type='solid', fgColor=STATUS_COLORS.get(status, 'FFFFFF'))
        border = _thin_border()

        values = [
            day['day_num'],
            day['date'],
            day['day_name'],
            status.replace('_', ' ').title(),
            STATUS_LABELS.get(status, '?'),
            day['leave_type'] or '',
            day['notes'] or '',
            '',
        ]

        for col_idx, val in enumerate(values, start=1):
            cell        = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill   = fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center' if col_idx <= 5 else 'left', vertical='center')
            if col_idx in (1, 5):
                cell.font = Font(bold=True, size=9)

        ws.row_dimensions[row_idx].height = 15

    # ── Column widths ──────────────────────────────────────
    col_widths = [6, 12, 6, 12, 10, 12, 30, 18]
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── Legend sheet ───────────────────────────────────────
    ws_legend = wb.create_sheet(title='Legend')
    ws_legend['A1'] = 'Code'
    ws_legend['B1'] = 'Meaning'
    ws_legend['A1'].font = Font(bold=True)
    ws_legend['B1'].font = Font(bold=True)
    legend_rows = [
        ('P',  'Present'),
        ('A',  'Absent'),
        ('H',  'Half Day'),
        ('L',  'Leave'),
        ('HO', 'Holiday'),
        ('WO', 'Week Off / Sunday'),
        ('—',  'Future (not yet)'),
    ]
    for i, (code, label) in enumerate(legend_rows, start=2):
        ws_legend.cell(row=i, column=1, value=code)
        ws_legend.cell(row=i, column=2, value=label)

    return wb


def export_html_pdf(calendar_data):
    """Returns an HTML string suitable for browser printing / PDF save."""
    emp_name   = calendar_data['employee_name']
    emp_code   = calendar_data['employee_code']
    month_name = calendar_data['month_name']
    year       = calendar_data['year']
    days       = calendar_data['days']
    summary    = calendar_data['summary']

    STATUS_CSS = {
        'present':  'background:#D1FAE5;color:#065F46',
        'absent':   'background:#FEE2E2;color:#991B1B',
        'half_day': 'background:#FEF9C3;color:#78350F',
        'leave':    'background:#DBEAFE;color:#1E40AF',
        'holiday':  'background:#EDE9FE;color:#5B21B6',
        'week_off': 'background:#F3F4F6;color:#6B7280',
        'future':   'background:#FFFFFF;color:#9CA3AF',
    }

    rows_html = ''
    for day in days:
        st    = day['status']
        style = STATUS_CSS.get(st, '')
        label = STATUS_LABELS.get(st, '?')
        rows_html += (
            f'<tr style="{style}">'
            f'<td style="text-align:center;font-weight:600">{day["day_num"]}</td>'
            f'<td>{day["date"]}</td>'
            f'<td style="text-align:center">{day["day_name"]}</td>'
            f'<td style="text-align:center">{st.replace("_"," ").title()}</td>'
            f'<td style="text-align:center;font-weight:700">{label}</td>'
            f'<td>{day["leave_type"] or ""}</td>'
            f'<td>{day["notes"] or ""}</td>'
            f'</tr>'
        )

    return f"""<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8">
<title>Attendance — {emp_name} — {month_name} {year}</title>
<style>
  body  {{ font-family: Arial, sans-serif; font-size: 12px; margin: 24px; color: #111; }}
  h2    {{ text-align: center; margin-bottom: 4px; }}
  .sub  {{ text-align: center; color: #6B7280; font-size: 11px; margin-bottom: 12px; }}
  table {{ width: 100%; border-collapse: collapse; }}
  th, td {{ border: 1px solid #D1D5DB; padding: 5px 8px; }}
  th    {{ background: #1F2937; color: #fff; text-align: center; }}
  .summary {{ display: flex; gap: 16px; flex-wrap: wrap; margin-bottom: 12px; font-size: 11px; }}
  .chip {{ padding: 3px 8px; border-radius: 12px; font-weight: 600; }}
  @media print {{ body {{ margin: 8px; }} }}
</style>
</head>
<body>
  <h2>Attendance Report</h2>
  <div class="sub">{emp_name} &nbsp;·&nbsp; {emp_code} &nbsp;·&nbsp; {month_name} {year}</div>
  <div class="summary">
    <span class="chip" style="background:#D1FAE5;color:#065F46">Present: {summary['present']}</span>
    <span class="chip" style="background:#FEE2E2;color:#991B1B">Absent: {summary['absent']}</span>
    <span class="chip" style="background:#FEF9C3;color:#78350F">Half Day: {summary['half_day']}</span>
    <span class="chip" style="background:#DBEAFE;color:#1E40AF">Leave: {summary['leave']}</span>
    <span class="chip" style="background:#EDE9FE;color:#5B21B6">Holiday: {summary['holiday']}</span>
    <span class="chip" style="background:#F3F4F6;color:#6B7280">Week Off: {summary['week_off']}</span>
    <span class="chip" style="background:#E5E7EB;color:#374151">Paid: {summary['paid_days']}/{summary['working_days']}</span>
  </div>
  <table>
    <thead>
      <tr>
        <th>#</th><th>Date</th><th>Day</th><th>Status</th><th>Code</th><th>Leave Type</th><th>Notes</th>
      </tr>
    </thead>
    <tbody>
      {rows_html}
    </tbody>
  </table>
  <p style="font-size:10px;color:#9CA3AF;margin-top:12px;text-align:right">Generated on print</p>
</body>
</html>"""
